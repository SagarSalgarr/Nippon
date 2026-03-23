"""
build_intent_model.py
=====================
Data-driven intent model pipeline. Zero hardcoding — everything driven by:
  intent_config.json  →  excel → clean data → DistilBERT fine-tune → ONNX INT8

To retrain when Excel data changes:
    python3 build_intent_model.py

To retrain with a different config:
    python3 build_intent_model.py --config my_config.json
"""

import argparse, json, re, pathlib, warnings, unicodedata
import numpy as np
import openpyxl
from collections import Counter, defaultdict

warnings.filterwarnings("ignore")


# ── Config loader ─────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    with open(path) as f:
        cfg = json.load(f)
    # Parse floats that JSON stores as strings (e.g. 3e-5)
    train = cfg.get("training", {})
    if isinstance(train.get("learning_rate"), str):
        train["learning_rate"] = float(train["learning_rate"])
    return cfg


# ── Text utilities ────────────────────────────────────────────────────────────

def clean_text(t: str) -> str:
    t = unicodedata.normalize("NFC", t.strip())
    t = re.sub(r"\s+", " ", t)
    return t

def is_mostly_devanagari(text: str, threshold: float) -> bool:
    deva = sum(1 for c in text if "\u0900" <= c <= "\u097F")
    return deva / max(len(text), 1) > threshold

def matches_any_skip_pattern(text: str, patterns: list[str]) -> bool:
    for p in patterns:
        if re.match(p, text):
            return True
    return False

def apply_text_overrides(text: str, overrides: list[dict]) -> str | None:
    """Returns intent if a text-level rule matches, else None."""
    for rule in overrides:
        if re.search(rule["pattern"], text):
            return rule.get("intent")  # can be null → skip
    return None  # no match


# ── Excel extraction ──────────────────────────────────────────────────────────

def extract_from_excel(cfg: dict) -> list[dict]:
    """
    Reads the Excel file and returns a list of dicts:
      {text, intent, sheet, flow}
    Intent is resolved by:
      1. text_level_overrides (regex, highest priority)
      2. flow_to_intent map
      3. None (skip)
    """
    xl_path = cfg["excel_path"]
    skip_sheets = set(cfg.get("skip_sheets", []))
    sys_markers = [m.lower() for m in cfg.get("system_response_markers", ["respond"])]
    skip_patterns = cfg.get("skip_text_patterns", [])
    deva_thresh = cfg.get("mostly_hindi_threshold", 0.3)
    flow_map = cfg.get("flow_to_intent", {})
    text_rules = cfg.get("text_level_overrides", [])

    wb = openpyxl.load_workbook(xl_path)
    samples = []

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # "Common" sheet has different column layout: col0=English, col1=Hindi
        is_common = sheet_name == "Common"
        if is_common:
            for row in rows[1:]:
                en = clean_text(str(row[0])) if row[0] else ""
                if not en or matches_any_skip_pattern(en, skip_patterns):
                    continue
                if is_mostly_devanagari(en, deva_thresh):
                    continue
                # Try text-level overrides; skip if no match (Common has no flow context)
                intent = apply_text_overrides(en, text_rules)
                if intent is None:
                    continue
                samples.append({"text": en, "intent": intent,
                                 "sheet": sheet_name, "flow": "common"})
            continue

        # Standard sheet: col0=Flow, col1=Screen, col2=ScreenName, col3=CardType, col4=English, col5=Hindi
        last_flow = sheet_name

        for row in rows[1:]:
            if not any(row):
                continue
            flow      = clean_text(str(row[0])) if row[0] else last_flow
            card_type = clean_text(str(row[3])) if len(row) > 3 and row[3] else ""
            english   = clean_text(str(row[4])) if len(row) > 4 and row[4] else ""

            if row[0]:
                last_flow = flow

            # Skip system responses
            if any(m in card_type.lower() for m in sys_markers):
                continue

            # Skip empty / header / pattern matches
            if not english or len(english) < 2:
                continue
            if matches_any_skip_pattern(english, skip_patterns):
                continue

            # Skip Devanagari-dominant strings (they're Indic, not English)
            if is_mostly_devanagari(english, deva_thresh):
                continue

            # Resolve intent: text-level overrides first
            intent = apply_text_overrides(english, text_rules)

            # Then flow map
            if intent is None:
                intent = flow_map.get(last_flow)

            # None means skip; explicit null in config also means skip
            if intent is None:
                continue

            samples.append({
                "text":   english,
                "intent": intent,
                "sheet":  sheet_name,
                "flow":   last_flow,
            })

    return samples


# ── Dataset builder ───────────────────────────────────────────────────────────

def build_dataset(cfg: dict) -> list[dict]:
    print("Extracting utterances from Excel...")
    raw = extract_from_excel(cfg)

    # Deduplicate
    seen = set()
    dataset = []
    for s in raw:
        if s["intent"] is None:
            continue
        key = s["text"].lower()
        if key not in seen:
            seen.add(key)
            dataset.append({"text": s["text"], "intent": s["intent"]})

    print(f"\nTotal unique samples: {len(dataset)}")
    dist = Counter(d["intent"] for d in dataset)
    print("Intent distribution:")
    for intent, cnt in sorted(dist.items(), key=lambda x: -x[1]):
        print(f"  {cnt:4d}  {intent}")
    print()

    # Warn about intents with very few samples
    for intent, cnt in dist.items():
        if cnt < 5:
            print(f"  ⚠ WARNING: '{intent}' has only {cnt} sample(s) — accuracy may be low")

    return dataset


# ── DistilBERT fine-tune ──────────────────────────────────────────────────────

def train_model(dataset: list[dict], cfg: dict, out_dir: pathlib.Path):
    import torch
    from transformers import (
        DistilBertTokenizerFast,
        DistilBertForSequenceClassification,
        TrainingArguments, Trainer,
    )
    from torch.utils.data import Dataset as TorchDataset
    from sklearn.model_selection import train_test_split
    from sklearn.preprocessing import LabelEncoder
    from sklearn.metrics import classification_report

    tcfg = cfg.get("training", {})
    epochs    = tcfg.get("epochs", 8)
    bs        = tcfg.get("batch_size", 32)
    lr        = float(tcfg.get("learning_rate", 3e-5))
    max_len   = tcfg.get("max_seq_len", 64)
    val_split = tcfg.get("val_split", 0.15)
    seed      = tcfg.get("seed", 42)
    base_model = cfg.get("model_base", "distilbert-base-uncased")

    le = LabelEncoder()
    texts  = [d["text"] for d in dataset]
    labels = le.fit_transform([d["intent"] for d in dataset])
    id2label = {i: str(l) for i, l in enumerate(le.classes_)}
    label2id = {l: i for i, l in id2label.items()}

    print(f"Model: {base_model}")
    print(f"Classes ({len(le.classes_)}): {list(le.classes_)}")

    X_tr, X_val, y_tr, y_val = train_test_split(
        texts, labels, test_size=val_split, random_state=seed, stratify=labels
    )
    print(f"Train: {len(X_tr)}  Val: {len(X_val)}")

    tok = DistilBertTokenizerFast.from_pretrained(base_model)

    class IntentDS(TorchDataset):
        def __init__(self, texts, labels):
            enc = tok(texts, truncation=True, padding="max_length",
                      max_length=max_len, return_tensors="pt")
            self.ids   = enc["input_ids"]
            self.masks = enc["attention_mask"]
            self.labs  = torch.tensor(labels, dtype=torch.long)
        def __len__(self): return len(self.labs)
        def __getitem__(self, i):
            return {"input_ids": self.ids[i], "attention_mask": self.masks[i],
                    "labels": self.labs[i]}

    tr_ds  = IntentDS(X_tr,  y_tr)
    val_ds = IntentDS(X_val, y_val)

    model = DistilBertForSequenceClassification.from_pretrained(
        base_model, num_labels=len(le.classes_),
        id2label=id2label, label2id=label2id,
    )

    def metrics(p):
        preds = np.argmax(p.predictions, axis=-1)
        return {"accuracy": float((preds == p.label_ids).mean())}

    ck_dir = str(out_dir / "checkpoints")
    args = TrainingArguments(
        output_dir=ck_dir,
        num_train_epochs=epochs,
        per_device_train_batch_size=bs,
        per_device_eval_batch_size=bs * 2,
        learning_rate=lr,
        warmup_ratio=0.1,
        weight_decay=0.01,
        eval_strategy="epoch",
        save_strategy="epoch",
        load_best_model_at_end=True,
        metric_for_best_model="accuracy",
        logging_steps=20,
        fp16=torch.cuda.is_available(),
        report_to="none",
        seed=seed,
    )

    trainer = Trainer(
        model=model, args=args,
        train_dataset=tr_ds, eval_dataset=val_ds,
        compute_metrics=metrics,
    )

    print("\nFine-tuning DistilBERT...")
    trainer.train()
    print("\nTraining complete. Evaluating...")

    preds = trainer.predict(val_ds)
    y_pred = np.argmax(preds.predictions, axis=-1)
    print("\n=== Validation Classification Report ===")
    print(classification_report(y_val, y_pred,
                                target_names=le.classes_, zero_division=0))

    return model, tok, le, id2label


# ── ONNX export + INT8 quantize ───────────────────────────────────────────────

def export_onnx(model, tokenizer, id2label: dict, cfg: dict, out_dir: pathlib.Path):
    import torch
    from onnxruntime.quantization import quantize_dynamic, QuantType

    out_dir.mkdir(parents=True, exist_ok=True)
    max_len  = cfg.get("training", {}).get("max_seq_len", 64)
    fp_path  = out_dir / "intent_model.onnx"
    int8_path = out_dir / "intent_model_int8.onnx"

    model.eval()
    dummy = tokenizer(["I want to create a gullak"], return_tensors="pt",
                      padding="max_length", max_length=max_len, truncation=True)

    print("\nExporting to ONNX (fp32, TorchScript path)...")
    model.eval()
    with torch.no_grad():
        torch.onnx.export(
            model,
            (dummy["input_ids"], dummy["attention_mask"]),
            str(fp_path),
            input_names=["input_ids", "attention_mask"],
            output_names=["logits"],
            dynamic_axes={
                "input_ids":      {0: "batch", 1: "seq"},
                "attention_mask": {0: "batch", 1: "seq"},
                "logits":         {0: "batch"},
            },
            opset_version=14,
            dynamo=False,   # use legacy TorchScript exporter (compatible with quantize_dynamic)
        )
    print(f"  {fp_path}  ({fp_path.stat().st_size/1e6:.1f} MB)")

    print("Quantizing to INT8...")
    quantize_dynamic(str(fp_path), str(int8_path), weight_type=QuantType.QInt8)
    print(f"  {int8_path}  ({int8_path.stat().st_size/1e6:.1f} MB)")

    # Save tokenizer
    tok_dir = out_dir / "tokenizer"
    tok_dir.mkdir(exist_ok=True)
    tokenizer.save_pretrained(str(tok_dir))

    # Save labels
    (out_dir / "intent_labels.json").write_text(
        json.dumps(id2label, indent=2, ensure_ascii=False))

    # Save human-readable descriptions
    (out_dir / "intent_descriptions.json").write_text(
        json.dumps(cfg.get("intent_descriptions", {}), indent=2, ensure_ascii=False))

    print(f"  Tokenizer → {tok_dir}")
    print(f"  Labels    → {out_dir/'intent_labels.json'}")

    return int8_path


# ── Inference test ────────────────────────────────────────────────────────────

def test_inference(model_path: pathlib.Path, tok_dir: pathlib.Path, id2label: dict):
    import onnxruntime as ort
    from transformers import DistilBertTokenizerFast

    print("\n=== Live Inference Test ===")
    sess  = ort.InferenceSession(str(model_path))
    tok   = DistilBertTokenizerFast.from_pretrained(str(tok_dir))
    shape1 = sess.get_inputs()[0].shape[1]
    max_len = shape1 if isinstance(shape1, int) and shape1 > 0 else 64

    def predict(text: str) -> tuple[str, float]:
        enc = tok(text, return_tensors="np", padding="max_length",
                  max_length=max_len, truncation=True)
        logits = sess.run(["logits"], {
            "input_ids":      enc["input_ids"].astype(np.int64),
            "attention_mask": enc["attention_mask"].astype(np.int64),
        })[0][0]
        exps = np.exp(logits - logits.max())
        probs = exps / exps.sum()
        idx  = int(np.argmax(probs))
        conf = float(probs[idx] * 100)
        label = id2label.get(idx) or id2label.get(str(idx)) or "unknown"
        return label, conf

    tests = [
        "I want to create a gullak",
        "What is my balance",
        "Yes, proceed",
        "SIP monthly",
        "UPI",
        "Budhape ka Sahara Gullak",
        "Monthly",
        "One thousand rupees",
        "Add nominee",
        "Login",
        "Enter OTP",
        "KYC verification",
        "Bank mandate setup",
        "Go back",
        "Namaste",
        "I need to check how much money I have in my portfolio",
        "Please help me open a retirement savings fund",
        "I want to invest money every month automatically",
        "Lumpsum investment of five thousand rupees",
        "Show me the gullak details",
    ]

    for text in tests:
        intent, conf = predict(text)
        print(f"  [{intent:25s}] {conf:5.1f}%  '{text}'")

    print()
    # Interactive mode
    print("Enter text to test (empty to quit):")
    while True:
        try:
            t = input("  > ").strip()
        except (EOFError, KeyboardInterrupt):
            break
        if not t:
            break
        intent, conf = predict(t)
        print(f"    → [{intent}]  {conf:.1f}% confidence")


# ── Packaging for S3 ─────────────────────────────────────────────────────────

def package_for_s3(out_dir: pathlib.Path, dist_dir: pathlib.Path):
    import zipfile, hashlib, shutil

    dist_dir.mkdir(parents=True, exist_ok=True)
    zip_name = dist_dir / "intent-distilbert-int8-v1.zip"

    print(f"\nPackaging for S3: {zip_name}")
    with zipfile.ZipFile(zip_name, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in [
            out_dir / "intent_model_int8.onnx",
            out_dir / "intent_labels.json",
            out_dir / "intent_descriptions.json",
        ]:
            if f.exists():
                zf.write(f, f.name)
        # Tokenizer files
        tok_dir = out_dir / "tokenizer"
        for f in tok_dir.rglob("*"):
            if f.is_file():
                zf.write(f, f"tokenizer/{f.name}")

    # SHA-256
    sha = hashlib.sha256(zip_name.read_bytes()).hexdigest()
    (dist_dir / "intent-distilbert-int8-v1.zip.sha256").write_text(sha)

    size_mb = zip_name.stat().st_size / 1e6
    print(f"  {zip_name} ({size_mb:.1f} MB)")
    print(f"  SHA256: {sha}")
    print()
    print("Next steps:")
    print("  1. Update manifest.json to add intent model entry")
    print("  2. python3 step4_upload.py   ← pushes zip to S3")
    print("  3. Sync CloudFront CDN")

    return zip_name, sha


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="intent_config.json",
                        help="Path to config JSON (default: intent_config.json)")
    parser.add_argument("--skip-training", action="store_true",
                        help="Skip training, only export/test existing checkpoint")
    parser.add_argument("--no-package", action="store_true",
                        help="Skip S3 packaging step")
    args = parser.parse_args()

    cfg     = load_config(args.config)
    out_dir = pathlib.Path(cfg["output_dir"])
    dist_dir = pathlib.Path("dist")

    dataset = build_dataset(cfg)

    if not args.skip_training:
        model, tok, le, id2label = train_model(dataset, cfg, out_dir)
        int8_path = export_onnx(model, tok, id2label, cfg, out_dir)
    else:
        from transformers import DistilBertForSequenceClassification, DistilBertTokenizerFast
        import glob as _glob
        # Find best checkpoint (highest step number)
        ck_dir = out_dir / "checkpoints"
        n_classes = len(build_dataset(cfg))  # re-use to know expected class count via label set
        # Actually just count from dataset directly
        expected_labels = len({d["intent"] for d in build_dataset(cfg)})
        all_ck = sorted(_glob.glob(str(ck_dir / "checkpoint-*")),
                        key=lambda p: int(p.rsplit("-", 1)[-1]))
        if not all_ck:
            raise FileNotFoundError(f"No checkpoints found in {ck_dir}")
        # Prefer checkpoint matching expected label count, pick highest step
        matching = [c for c in all_ck if len(json.loads(
            open(f"{c}/config.json").read()).get("id2label", {})) == expected_labels]
        best_ck = (matching or all_ck)[-1]
        print(f"Loading checkpoint: {best_ck}")
        model = DistilBertForSequenceClassification.from_pretrained(best_ck)
        base_model = cfg.get("model_base", "distilbert-base-uncased")
        tok   = DistilBertTokenizerFast.from_pretrained(base_model)
        id2label = {int(k): v for k, v in model.config.id2label.items()}
        int8_path = export_onnx(model, tok, id2label, cfg, out_dir)

    test_inference(int8_path, out_dir / "tokenizer", id2label)

    if not args.no_package:
        package_for_s3(out_dir, dist_dir)

    print(f"\n✓ All done. Model: {int8_path}")


if __name__ == "__main__":
    main()
